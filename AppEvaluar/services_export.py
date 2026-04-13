import openpyxl
from io import BytesIO
from django.db.models import Q
from AppGestionUsuario.models import Profile
from AppTutoria.models import ProgresoEstudiante, Tema
from .services_metrics import get_classroom_performance_summary

def generar_excel_reporte_docente(grado=None, seccion=None, nombre=None, tema_id=None, fecha_inicio=None, fecha_fin=None):
    """
    Genera un archivo Excel (.xlsx) con el reporte analítico del aula.
    """
    wb = openpyxl.Workbook()
    
    # 1. Hoja: Resumen de Aula
    ws_resumen = wb.active
    ws_resumen.title = "Resumen de Aula"
    
    summary = get_classroom_performance_summary(
        grado=grado, seccion=seccion, 
        fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, 
        tema_id=tema_id
    )
    
    ws_resumen.append(["Métrica", "Valor"])
    ws_resumen.append(["Total Estudiantes", summary['total_estudiantes']])
    ws_resumen.append(["Precisión Promedio (%)", f"{summary['precision_promedio']}%"])
    ws_resumen.append(["XP Promedio", summary['puntos_promedio']])
    
    # Agregar desglose por tema si existe
    if summary['desempeno_por_tema']:
        ws_resumen.append([])
        ws_resumen.append(["Dominio por Tema", "Porcentaje (%)"])
        for tema, valor in summary['desempeno_por_tema'].items():
            ws_resumen.append([tema, f"{valor}%"])

    # 2. Hoja: Listado de Estudiantes
    ws_estudiantes = wb.create_sheet(title="Listado de Estudiantes")
    headers_est = ["Nombre", "Usuario", "Grado/Sección", "XP", "Nivel", "Precisión", "Insignias"]
    ws_estudiantes.append(headers_est)
    
    # Obtener estudiantes filtrados
    est_queryset = Profile.objects.filter(rol='Estudiante').select_related('user').prefetch_related('user__metricas', 'logros')
    if grado:
        est_queryset = est_queryset.filter(grado=grado)
    if seccion:
        est_queryset = est_queryset.filter(seccion=seccion)
    if nombre:
        est_queryset = est_queryset.filter(
            Q(user__username__icontains=nombre) | 
            Q(nombres__icontains=nombre) | 
            Q(apellidos__icontains=nombre)
        )
    
    for est in est_queryset:
        try:
            metricas = est.user.metricas.first()
            precision = f"{metricas.precision_general:.1f}%" if metricas else "Sin datos"
        except (AttributeError, Exception):
            precision = "Sin datos"
        
        ws_estudiantes.append([
            f"{est.nombres} {est.apellidos}",
            est.user.username,
            f"{est.grado} {est.seccion}",
            est.puntos_acumulados,
            est.nivel_estudiante,
            precision,
            est.logros.count()
        ])

    # 3. Hoja: Actividad Reciente
    ws_actividad = wb.create_sheet(title="Actividad Reciente")
    headers_act = ["Fecha/Hora", "Estudiante", "Aula", "Tema", "Actividad"]
    ws_actividad.append(headers_act)
    
    # Obtener actividad filtrada
    prog_queryset = ProgresoEstudiante.objects.all().select_related('usuario', 'tema')
    if grado:
        prog_queryset = prog_queryset.filter(grado=grado)
    if seccion:
        prog_queryset = prog_queryset.filter(seccion=seccion)
    if nombre:
        prog_queryset = prog_queryset.filter(
            Q(usuario__username__icontains=nombre) | 
            Q(usuario__first_name__icontains=nombre) | 
            Q(usuario__last_name__icontains=nombre)
        )
    if tema_id:
        prog_queryset = prog_queryset.filter(tema_id=tema_id)
    if fecha_inicio:
        prog_queryset = prog_queryset.filter(fecha_registro__date__gte=fecha_inicio)
    if fecha_fin:
        prog_queryset = prog_queryset.filter(fecha_registro__date__lte=fecha_fin)
        
    # Limitar a los últimos 100 registros para no sobrecargar el Excel
    actividades = prog_queryset.order_by('-fecha_registro')[:100]
    
    for p in actividades:
        ws_actividad.append([
            p.fecha_registro.strftime("%d/%m/%Y %H:%M"),
            p.usuario.get_full_name() or p.usuario.username,
            f"{p.grado} {p.seccion}",
            p.tema.nombre if p.tema else "N/A",
            p.tipo_actividad
        ])

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
