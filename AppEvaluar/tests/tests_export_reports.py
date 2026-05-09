from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth.models import User
from AppGestionUsuario.models import Profile
from AppTutoria.models import Tema, ProgresoEstudiante
import openpyxl
from io import BytesIO

class ExportReportsTests(TestCase):
    def setUp(self):
        self.client = Client()
        
        # Crear usuarios con diferentes roles
        self.docente_user = User.objects.create_user(username='docente', password='password123')
        self.docente_profile = Profile.objects.create(user=self.docente_user, rol='Docente', nombres='Docente', apellidos='Prueba')
        
        self.estudiante_user = User.objects.create_user(username='estudiante', password='password123')
        self.estudiante_profile = Profile.objects.create(
            user=self.estudiante_user, 
            rol='Estudiante', 
            nombres='Estudiante', 
            apellidos='Prueba',
            grado='2do',
            seccion='A'
        )
        
        # Crear datos de prueba
        self.tema = Tema.objects.create(nombre="Triángulos", slug="triangulos")
        ProgresoEstudiante.objects.create(
            usuario=self.estudiante_user,
            tema=self.tema,
            tipo_actividad='Ejercicio',
            grado='2do',
            seccion='A'
        )
        
        self.export_url = reverse('evaluar:exportar_reporte_excel')

    def test_export_access_restricted_to_docentes(self):
        """La URL de exportación solo debe ser accesible para docentes."""
        # Intento como estudiante (debe fallar con 403)
        self.client.login(username='estudiante', password='password123')
        response = self.client.get(self.export_url)
        self.assertEqual(response.status_code, 403)
        
        # Intento como docente (debe ser exitoso 200)
        self.client.login(username='docente', password='password123')
        response = self.client.get(self.export_url)
        self.assertEqual(response.status_code, 200)

    def test_export_excel_format_and_headers(self):
        """Verificar que se genera un archivo .xlsx con las hojas y encabezados correctos."""
        self.client.login(username='docente', password='password123')
        response = self.client.get(self.export_url)
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Abrir el Excel generado en memoria
        wb = openpyxl.load_workbook(BytesIO(response.content))
        
        # Verificar existencia de las tres hojas solicitadas en la especificación
        self.assertIn("Resumen de Aula", wb.sheetnames)
        self.assertIn("Listado de Estudiantes", wb.sheetnames)
        self.assertIn("Actividad Reciente", wb.sheetnames)
        
        # Verificar encabezados en la hoja de Estudiantes (Hoja 2)
        sheet_est = wb["Listado de Estudiantes"]
        expected_headers = ["Nombre", "Usuario", "Grado/Sección", "XP", "Nivel", "Precisión", "Insignias"]
        actual_headers = [cell.value for cell in sheet_est[1]]
        for header in expected_headers:
            self.assertIn(header, actual_headers)

    def test_export_respects_filters(self):
        """Verificar que la exportación respete los filtros de grado y sección."""
        # Crear otro estudiante en una sección diferente
        other_user = User.objects.create_user(username='estudiante_b', password='password123')
        Profile.objects.create(
            user=other_user, 
            rol='Estudiante', 
            nombres='Estudiante', 
            apellidos='B',
            grado='2do',
            seccion='B'
        )
        ProgresoEstudiante.objects.create(
            usuario=other_user,
            tema=self.tema,
            tipo_actividad='Ejercicio',
            grado='2do',
            seccion='B'
        )
        
        self.client.login(username='docente', password='password123')
        
        # Filtrar por Sección A
        response = self.client.get(self.export_url, {'seccion': 'A'})
        wb = openpyxl.load_workbook(BytesIO(response.content))
        sheet_act = wb["Actividad Reciente"]
        
        # Contar filas de datos (excluyendo encabezado)
        data_rows = list(sheet_act.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(data_rows), 1)
        self.assertEqual(data_rows[0][2], '2do A') # Columna Aula/Sección (Índice 2)
        
        # Filtrar por Sección B
        response = self.client.get(self.export_url, {'seccion': 'B'})
        wb = openpyxl.load_workbook(BytesIO(response.content))
        sheet_act = wb["Actividad Reciente"]
        data_rows = list(sheet_act.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(data_rows), 1)
        self.assertEqual(data_rows[0][2], '2do B')
